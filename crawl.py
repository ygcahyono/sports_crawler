#!/usr/bin/env python3
import argparse
import json
import math
import os
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from playwright.sync_api import sync_playwright

load_dotenv()

SITE_URL = "https://kuyy.id"
API_URL = "https://kuyy.app/api/events"
WIB = timezone(timedelta(hours=7))

TYPE_KEYWORDS = [("single", "Single"), ("double", "Double"), ("coach", "Coaching")]


def classify_type(name: str) -> str:
    if not name:
        return ""
    lower = name.lower()
    return ", ".join(label for kw, label in TYPE_KEYWORDS if kw in lower)


def load_locations() -> dict:
    locations = {}
    prefix = "LOC_"
    for key, val in os.environ.items():
        if key.startswith(prefix):
            name = key[len(prefix):].lower()
            try:
                lat, lon = val.split(",")
                locations[name] = (float(lat.strip()), float(lon.strip()))
            except ValueError:
                pass
    return locations


def parse_args():
    parser = argparse.ArgumentParser(description="Crawl tennis activities from kuyy.id")
    parser.add_argument("--date", required=True, help="Date in YYYY-MM-DD format")
    parser.add_argument("--start-time", default="00:00", help="Start time filter HH:MM in WIB (default: 00:00)")
    parser.add_argument("--end-time", default="23:59", help="End time filter HH:MM in WIB (default: 23:59)")
    parser.add_argument("--output", "-o", help="Output Excel file path (default: output/kuyy_tennis_<date>.xlsx)")
    parser.add_argument("--sort-by", choices=["time", "distance", "price"], default="time", help="Sort results by time, distance, or price (default: time)")
    parser.add_argument("--from", dest="location", default="home", help="Location name for distance sort/column, e.g. home, work, trc, kemang (default: home)")
    parser.add_argument("--no-headless", action="store_true", help="Show browser window")
    parser.add_argument("--otp", help="OTP code for login (skips interactive prompt)")
    parser.add_argument("--send-otp", action="store_true", help="Only send OTP email, don't crawl yet")
    return parser.parse_args()


def send_otp(page, email: str):
    page.goto(f"{SITE_URL}/login", wait_until="networkidle")
    page.fill('input[type="email"]', email)
    page.click('button[type="submit"]')
    page.wait_for_selector('input[placeholder="Enter OTP"]', timeout=15000)
    print(f"OTP sent to {email}. Check your email.")


def login(page, email: str, otp: str = None):
    print("Logging in...")
    page.goto(f"{SITE_URL}/login", wait_until="networkidle")
    page.fill('input[type="email"]', email)
    page.click('button[type="submit"]')
    page.wait_for_selector('input[placeholder="Enter OTP"]', timeout=15000)
    print(f"OTP sent to {email}.")
    if not otp:
        otp = input("Enter OTP code: ").strip()
    page.fill('input[placeholder="Enter OTP"]', otp)
    page.click('button[type="submit"]')
    page.wait_for_url(lambda url: "/login" not in url, timeout=30000)
    print("Login successful.")


def fetch_all_activities(page, date: str) -> list:
    url = f"{API_URL}?category=tennis&date={date}&limit=1000"
    resp = page.request.get(url)
    if resp.status != 200:
        print(f"API error: {resp.status}")
        return []

    data = resp.json().get("data") or []
    print(f"  Fetched {len(data)} activities from API.")
    # API ignores the category param and returns all sports; filter locally.
    tennis = [a for a in data if "tennis" in (a.get("categories") or [])]
    print(f"  Kept {len(tennis)} tennis activities.")
    return tennis


def timestamp_to_wib(ts: str) -> str:
    dt = datetime.fromisoformat(ts.replace("Z", "+00:00"))
    return dt.astimezone(WIB).strftime("%H:%M")


def filter_by_time(activities: list, start_time: str, end_time: str) -> list:
    filtered = []
    for act in activities:
        ts = act.get("start_timestamp", "")
        if not ts:
            continue
        wib_time = timestamp_to_wib(ts)
        if start_time <= wib_time <= end_time:
            filtered.append(act)
    return filtered


def haversine_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    R = 6371.0
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat / 2) ** 2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon / 2) ** 2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def deduplicate(activities: list) -> list:
    seen = set()
    unique = []
    for act in activities:
        aid = act.get("id", "")
        if aid not in seen:
            seen.add(aid)
            unique.append(act)
    return unique


def activities_to_excel(activities: list, output_path: str, date: str, location: tuple = None, location_name: str = ""):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tennis Activities"

    has_distance = location is not None

    headers = [
        "Activity Name",
        "Host / Community",
        "Venue",
        "Address",
        "Date",
        "Start Time (WIB)",
        "End Time (WIB)",
        "Price (Rp)",
    ]
    if has_distance:
        headers.append(f"Dist from {location_name} (km)")
    headers.append("Type")
    headers.append("Link")

    header_fill = PatternFill(start_color="FF710B", end_color="FF710B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, act in enumerate(activities, 2):
        host_info = act.get("host_info") or {}
        host_name = host_info.get("name", "") if isinstance(host_info, dict) else ""

        values = [
            act.get("name", ""),
            host_name,
            act.get("location_label", ""),
            act.get("location", ""),
            date,
            timestamp_to_wib(act.get("start_timestamp", "")),
            timestamp_to_wib(act.get("end_timestamp", "")),
            act.get("price", 0),
        ]
        if has_distance:
            lat = act.get("latitude", 0) or 0
            lon = act.get("longitude", 0) or 0
            dist = round(haversine_km(location[0], location[1], lat, lon), 1) if lat and lon else None
            values.append(dist)
        values.append(classify_type(act.get("name", "")))
        values.append(f"{SITE_URL}/activities/{act.get('id', '')}")

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 4))

    col_widths = [45, 25, 30, 50, 12, 16, 16, 15]
    if has_distance:
        col_widths.append(14)
    col_widths.append(18)
    col_widths.append(55)
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    ws.auto_filter.ref = ws.dimensions

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Saved {len(activities)} activities to {output_path}")


def get_session_path() -> Path:
    return Path(__file__).parent / ".session"


def main():
    args = parse_args()

    email = os.getenv("KUYY_EMAIL")
    if not email:
        print("Error: Set KUYY_EMAIL in .env file.")
        print("See .env.example for reference.")
        sys.exit(1)

    headless = not args.no_headless
    session_path = get_session_path()

    if args.send_otp:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=headless)
            page = browser.new_context().new_page()
            send_otp(page, email)
            browser.close()
        print("Now re-run with --otp <code> to complete login and crawl.")
        return

    output = args.output or f"output/kuyy_tennis_{args.date}.xlsx"

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=headless)

        if session_path.exists() and not args.otp:
            context = browser.new_context(storage_state=str(session_path))
            print("Using saved session.")
        else:
            context = browser.new_context()

        page = context.new_page()

        # Load the site to establish cookies / check session
        page.goto(f"{SITE_URL}/activities/tennis", wait_until="networkidle")
        page.wait_for_timeout(2000)

        if "/login" in page.url:
            if not args.otp:
                print("Session expired or not found.")
                print("Step 1: Run with --send-otp to receive the OTP email.")
                print("Step 2: Re-run with --otp <code> to login and crawl.")
                browser.close()
                sys.exit(1)
            login(page, email, args.otp)
            context.storage_state(path=str(session_path))
            print("Session saved for future runs.")

        print(f"Fetching tennis activities for {args.date}...")
        activities = fetch_all_activities(page, args.date)
        browser.close()

    activities = deduplicate(activities)
    print(f"Found {len(activities)} total activities.")

    activities = filter_by_time(activities, args.start_time, args.end_time)
    print(f"After time filter ({args.start_time}-{args.end_time} WIB): {len(activities)} activities.")

    if not activities:
        print("No activities match the time filter.")
        sys.exit(0)

    locations = load_locations()
    loc_name = args.location.lower()
    loc_coords = locations.get(loc_name)

    if args.sort_by == "distance":
        if not loc_coords:
            available = ", ".join(sorted(locations.keys())) or "(none configured)"
            print(f"Error: Location '{loc_name}' not found in .env.")
            print(f"Available locations: {available}")
            print(f"Add LOC_{loc_name.upper()}=lat,lon to your .env file.")
            sys.exit(1)
        activities.sort(key=lambda a: haversine_km(loc_coords[0], loc_coords[1], a.get("latitude", 0) or 0, a.get("longitude", 0) or 0))
    elif args.sort_by == "price":
        activities.sort(key=lambda a: a.get("price", 0))
    else:
        activities.sort(key=lambda a: timestamp_to_wib(a.get("start_timestamp", "")))

    activities_to_excel(activities, output, args.date, location=loc_coords, location_name=loc_name)


if __name__ == "__main__":
    main()
